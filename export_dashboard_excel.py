from __future__ import annotations

import json
import subprocess
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
PARTNERS_JS = ROOT / "partners.js"
ARTEFACTS_JS = ROOT / "artefacts.js"
OUTPUT_XLSX = ROOT / "rainmatter_partner_dashboard_export.xlsx"
PAYLOAD_SCRIPT = ROOT / "export_dashboard_payload.js"


def load_dashboard_payload() -> dict:
    result = subprocess.run(
        ["node", str(PAYLOAD_SCRIPT), str(PARTNERS_JS), str(ARTEFACTS_JS)],
        check=True,
        capture_output=True,
        text=True,
    )
    return json.loads(result.stdout)


def write_sheet(workbook: Workbook, title: str, rows: list[dict]) -> None:
    ws = workbook.create_sheet(title=title)
    if not rows:
        ws.append(["No data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill(fill_type="solid", fgColor="16796E")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws[get_column_letter(idx)]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 48)


def main() -> None:
    payload = load_dashboard_payload()

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    write_sheet(wb, "Summary", payload["summary"])
    write_sheet(wb, "Partners", payload["partners"])
    write_sheet(wb, "Artefacts Overview", payload["overview"])
    write_sheet(wb, "Tangible Artefacts", payload["tangible"])
    write_sheet(wb, "Intangible Artefacts", payload["intangible"])
    write_sheet(wb, "Sources", payload["sources"])

    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
